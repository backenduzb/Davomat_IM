from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_protect
from utils.upload import upload_document
from teachers.models import Class
from openpyxl import Workbook
from openpyxl.styles import Alignment
from students.models import Student
from utils.time import current_time
from django.conf import settings
import time
import os

def is_admin(user):
    return user.is_authenticated and user.is_staff

@csrf_protect
@login_required
@user_passes_test(is_admin)
@transaction.atomic
def export_data():
    now_caption = current_time()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"

    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"] = "Kelmagan o‘quvchilar ro‘yxati"

    headers = ["№", "F.I.SH", "SINF", "SANA", "SABAB"]
    widths = [3, 60, 7, 12, 60]

    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64 + col)].width = width

    students = Student.objects.exclude(status="Bor")

    row = 3
    for idx, student in enumerate(students, start=1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=student.full_name)
        ws.cell(row=row, column=3, value=str(student.class_type.class_name))
        ws.cell(row=row, column=4, value=now_caption)
        ws.cell(row=row, column=5, value=student.sababi or "-")
        row += 1

    file_name = os.path.join(settings.BASE_DIR, "Davomat.xlsx")
    
    try:
        os.remove(file_name)
    except FileNotFoundError:
        pass
    wb.save(file_name)
    time.sleep(3)
    upload_document(document_path=file_name)
    
    time.sleep(3)
    Student.objects.update(status="Bor", sababi="")
    Class.objects.update(this_updated=False)
