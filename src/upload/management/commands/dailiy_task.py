from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Alignment
from students.models import Student
from teachers.models import Class
from django.conf import settings
from upload.tg_uploader import upload_document
import os, time

class Command(BaseCommand):
    help = "Kelmagan o'quvchilarni Excel faylga eksport qiladi"

    def handle(self, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Davomat"

        ws.merge_cells("A1:E1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"] = "Kelmagan o‘quvchilar ro‘yxati"

        headers = ["№", "F.I.SH", "SINF", "SANA", "SABAB"]
        widths = [3, 60, 7, 12, 20]

        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[chr(64 + col)].width = width

        students = Student.objects.exclude(status="Bor")

        row_num = 3
        for idx, student in enumerate(students, start=1):
            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=student.full_name)
            ws.cell(row=row_num, column=3, value=str(student.class_type.class_name))
            ws.cell(row=row_num, column=4, value="")
            ws.cell(row=row_num, column=5, value=student.sababi or "-")
            row_num += 1

        file_name = os.path.join(settings.BASE_DIR, "Davomat.xlsx")

        if os.path.exists(file_name):
            os.remove(file_name)
        wb.save(file_name)

        time.sleep(5)
        upload_document(document=file_name)
        time.sleep(5)
        Student.objects.update(sababi=" ",status="Bor")
        Class.objects.update(this_updated=False)
        self.stdout.write(self.style.SUCCESS(f"{len(students)} ta kelmagan o‘quvchi '{file_name}' fayliga eksport qilindi ✅"))
