from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

wb = Workbook()

wb.title = "Datomat"

ws = wb.active 

ws.merge_cells("A1:E1")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws["A1"]="Dushanba"

ws.column_dimensions["A"].width = 3
ws["A2"] = "№"
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.column_dimensions["B"].width = 60
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws["B2"] = "F.I.SH"

ws.column_dimensions["C"].width = 7
ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
ws["C2"] = "SINF"

ws.column_dimensions["D"].width = 12
ws["D2"].alignment = Alignment(horizontal="center", vertical="center")
ws["D2"] ="SANA"

ws.column_dimensions["E"].width = 7
ws["E2"].alignment = Alignment(horizontal="center", vertical="center")
ws["E2"] = "SABAB"


if os.path.exists("Davomat.xlsx"):
    os.remove("Davomat.xlsx")
    wb.save("Davomat.xlsx")
else:
    wb.save("Davomat.xlsx")


